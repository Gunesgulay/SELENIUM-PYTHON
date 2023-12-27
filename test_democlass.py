from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as ec
from selenium.webdriver.common.action_chains import ActionChains
import pytest
import openpyxl
from constants import globalContants as c
        
class Test_democlass:
    def setup_method(self):
        chrome_driver_path = Service("/Users/gunesgulay/Downloads/chromedriver-mac-arm64/chromedriver")
        self.driver = webdriver.Chrome(service=chrome_driver_path)
        self.driver.get(c.BASE_URL)
        self.driver.maximize_window()

    def teardown_method(self):
        self.driver.quit()    

    def getData():
        excel = openpyxl.load_workbook(c.invalid_login_xlsx) 
        sayfa = excel["Sayfa1"]
        rows = sayfa.max_row
        data = []
        for i in range (2,rows+1):
            username = sayfa.cell(i,1).value
            password = sayfa.cell(i,2).value
            data.append((username, password))
             
        return data    
    
    @pytest.mark.parametrize("username, password", getData())
    def test_invalid_login(self, username, password):
        
        usernameInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,c.USERNAME_ID)))
        usernameInput.send_keys(username)

        passwordInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,c.PASSWORD_ID)))
        passwordInput.send_keys(password)

        loginButton = self.driver.find_element(By.ID,c.LOGIN_BUTTON_ID)
        loginButton.click()
        
        errorMessage = self.driver.find_element(By.XPATH,c.ERROR_MESSAGE_XPATH)
        assert errorMessage.text == c.USERNAME_PASSWORD_DONTMATCH


    @pytest.mark.skip   
    def test_valid_login(self):
        
        usernameInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID, c.USERNAME_ID)))
        usernameInput.send_keys("standard_user")

        passwordInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,c.PASSWORD_ID)))
        passwordInput.send_keys("secret_sauce")
        
        loginButton = self.driver.find_element(By.ID,c.LOGIN_BUTTON_ID)
        loginButton.click()
        
        addToCart = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.XPATH,"//*[@id='add-to-cart-test.allthethings()-t-shirt-(red)']")))
        actionChains = ActionChains(self.driver)
        actionChains.move_to_element(addToCart)
        actionChains.click()
        actionChains.perform()
        remove = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.XPATH,"//*[@id='remove-test.allthethings()-t-shirt-(red)']")))
        assert remove.text == "Remove"

    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    

    

    