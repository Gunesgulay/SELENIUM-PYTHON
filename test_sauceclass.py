from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as ec
import pytest
import openpyxl
from constants import globalContants as c

class Test_Sauceclass:
    def setup_method(self):
        chrome_driver_path = Service("/Users/gunesgulay/Downloads/chromedriver-mac-arm64/chromedriver")
        self.driver = webdriver.Chrome(service=chrome_driver_path)
        self.driver.get(c.BASE_URL)
        self.driver.maximize_window()

    def teardown_method(self):
        self.driver.quit()  

    def getData():
        excel = openpyxl.load_workbook("data/sauce_class.xlsx") 
        sayfa = excel["Sayfa1"]
        rows = sayfa.max_row
        data = []
        for i in range (2,rows+1):
            firstname = sayfa.cell(i, 1).value
            lastname = sayfa.cell(i, 2).value
            zipcode = sayfa.cell(i, 3).value
            data.append((firstname, lastname, zipcode))
             
        return data    

    def test_two_fields_empty(self):
        loginButton = self.driver.find_element(By.ID,c.LOGIN_BUTTON_ID)
        loginButton.click()

        errorMessage = self.driver.find_element(By.XPATH,c.ERROR_MESSAGE_XPATH)
        assert errorMessage.text == c.USERNAME_REQUIRED

    @pytest.mark.parametrize("username", [("standard_user")])
    def test_one_fields_empty(self, username):
        usernameInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,c.USERNAME_ID)))
        usernameInput.send_keys(username)

        loginButton = self.driver.find_element(By.ID,c.LOGIN_BUTTON_ID)
        loginButton.click()

        errorMessage = self.driver.find_element(By.XPATH,c.ERROR_MESSAGE_XPATH)
        assert errorMessage.text == c.PASSWORD_REQUIRED

    @pytest.mark.parametrize("username, password", [("locked_out_user", "secret_sauce")])    
    def test_invalid_login(self, username, password):
        usernameInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,c.USERNAME_ID)))
        usernameInput.send_keys(username)

        passwordInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,c.PASSWORD_ID)))
        passwordInput.send_keys(password)

        loginButton = self.driver.find_element(By.ID,c.LOGIN_BUTTON_ID)
        loginButton.click()

        errorMessage = self.driver.find_element(By.XPATH,c.ERROR_MESSAGE_XPATH)
        assert errorMessage.text == c.USERNAME_PASSWORD_DONTMATCH

    @pytest.mark.parametrize("username, password", [("standard_user", "secret_sauce")])    
    def test_product_count(self, username, password):
        usernameInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,c.USERNAME_ID)))
        usernameInput.send_keys(username)

        passwordInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,c.PASSWORD_ID)))
        passwordInput.send_keys(password)

        loginButton = self.driver.find_element(By.ID,c.LOGIN_BUTTON_ID)
        loginButton.click()

        listOfProducts = self.driver.find_elements(By.CLASS_NAME,c.LIST_PRODUCTS_CLASS)
        assert len(listOfProducts) == 6 

    def add_product_to_cart(self):
        usernameInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,c.USERNAME_ID)))
        usernameInput.send_keys("standard_user")

        passwordInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,c.PASSWORD_ID)))
        passwordInput.send_keys("secret_sauce")

        loginButton = self.driver.find_element(By.ID,c.LOGIN_BUTTON_ID)
        loginButton.click()

        addToCart = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,c.ADD_TO_CART_ID)))
        addToCart.click()

        goToCart = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.CSS_SELECTOR,c.GO_TO_CART_CSS)))
        goToCart.click()    

    @pytest.mark.parametrize("username, password", [("standard_user", "secret_sauce")])    
    def test_add_product_to_cart_controller(self, username, password):
        usernameInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,c.USERNAME_ID)))
        usernameInput.send_keys(username)

        passwordInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,c.PASSWORD_ID)))
        passwordInput.send_keys(password)

        loginButton = self.driver.find_element(By.ID,c.LOGIN_BUTTON_ID)
        loginButton.click()

        addToCart = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,c.ADD_TO_CART_ID)))
        addToCart.click()

        goToCart = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.CSS_SELECTOR,c.GO_TO_CART_CSS)))
        goToCart.click()

        productName = self.driver.find_element(By.CSS_SELECTOR, c.PRODUCT_NAME_CSS)
        assert productName.text == c.PRODUCT_NAME

    def test_checkout_all_fields_empty(self):
        self.add_product_to_cart()

        checkoutButton = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,c.CHECKOUT_BUTTON_ID)))
        checkoutButton.click()

        continueButton = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,c.CONTINUE_BUTTON_ID)))
        continueButton.click()

        errorMessageController = self.driver.find_element(By.CSS_SELECTOR,c.ERROR_MESSAGE_CONTROLLER_CSS)
        assert errorMessageController.text == c.ERROR_MESSAGE_CONTROLLER

    @pytest.mark.parametrize("firstname, lastname, zipcode", getData())   
    def test_successful_purchase(self, firstname, lastname, zipcode):
        self.add_product_to_cart()

        checkoutButton = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,c.CHECKOUT_BUTTON_ID)))
        checkoutButton.click() 

        enterFirstName = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,c.FIRST_NAME_ID)))
        enterFirstName.send_keys(firstname)

        enterLastName = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,c.LAST_NAME_ID)))
        enterLastName.send_keys(lastname)

        enterZipCode = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,c.ZIP_CODE_ID)))
        enterZipCode.send_keys(zipcode)

        continueButton = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,c.CONTINUE_BUTTON_ID)))
        continueButton.click()

        finishButton = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,c.FINISH_BUTTON_ID)))
        finishButton.click()

        messageController = self.driver.find_element(By.CSS_SELECTOR,c.ORDER_MESSAGE_CONTROLLER_CSS)
        assert messageController.text == c.ORDER_MESSAGE